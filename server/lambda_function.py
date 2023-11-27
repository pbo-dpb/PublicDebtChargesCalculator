from openpyxl import Workbook, load_workbook
import os
import json
import formulas
import shutil
import uuid
import sys


def validate_input(user_cell_values, requested_fields):

    # Iterate over user_cell_values and update the cell values in the workbook
    for cell in user_cell_values:
        
        # Validate that cell is a plausible cell reference
        if not cell.isalnum():
            raise ValueError("Invalid user input cell reference (must be alpha numeric).")

        # Validate that user_cell_values[cell] is a number
        if not isinstance(user_cell_values[cell], float) and not isinstance(user_cell_values[cell], int):
            raise ValueError("Invalid value for cell " + cell + " (must be an int or float).")
    
    i = 0
    for field in requested_fields:
        # Validate that cell is a plausible cell reference
        if not field.isalnum():
            raise ValueError("Invalid requested cell reference (must be alpha numeric) at index " + str(i) + ".")
        i += 1

    return None


def lambda_handler(event, context):
    
    try:
        payload = json.loads(event["body"])
        user_cell_values = payload["user_values"]
        requested_fields = payload["requested_fields"]

        if not isinstance(user_cell_values, dict) or not isinstance(requested_fields, list):
            raise ValueError("Invalid payload format")
        
        validate_input(user_cell_values, requested_fields)

    except (KeyError, ValueError) as e:
        return {
            'statusCode': 400,
            'body': json.dumps(str(e))
        }

    run_id = str(uuid.uuid4())
    temporary_file_path = '/tmp/' + run_id + '.xlsx'

    # Copy ./payload.xlsx to a temporary file with a random name in /tmp
    shutil.copy('payload.xlsx', temporary_file_path)


    wb = load_workbook(filename = temporary_file_path)
    # Get a list of all sheet names in the workbook
    
    machine_readable_sheet = wb['machine_readable']
    
    # Iterate over user_cell_values and update the cell values in the workbook
    for cell in user_cell_values:
        machine_readable_sheet[cell].value = float(user_cell_values[cell])

    # Save the workbook to the temporary file
    wb.save(temporary_file_path)


    # Execute workbook formulas
    xl_model = formulas.ExcelModel().loads(temporary_file_path).finish()
    xl_model.calculate()

    output_dir = "/tmp/" + run_id + "/"
    xl_model.write(dirpath=output_dir)


    # Find the only file in the output dir and retrieve it
    output_file_name = os.listdir(output_dir)[0]

    # Read the contents of the file
    output_file_path = output_dir + output_file_name


    ouput_wb = load_workbook(filename = output_file_path)

    # Iterate over requested_fields and retrieve the cell values from the workbook
    output_values = {}
    for field in requested_fields:
        output_values[field] = ouput_wb['MACHINE_READABLE'][field].value


    return {
        'statusCode': 200,
        'body': json.dumps(output_values)
    }
    
    # Clean up the temporary files
    os.remove(temporary_file_path)
    os.remove(output_file_path)

if __name__ == '__main__':
    print(lambda_handler({'body': '{"user_values":{"M3": 1000, "N3": 1000, "O3": 1000, "P3": 1000, "Q3": 1000}, "requested_fields": ["M4"]}'}, None))